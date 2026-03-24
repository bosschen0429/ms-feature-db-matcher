from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from .matcher import MatchCell

DNA_COLOR = "0000FF"
RNA_COLOR = "FF0000"


def _build_output_path(source_path: Path, output_dir: Path) -> Path:
    output_path = output_dir / f"{source_path.stem}_matched.xlsx"
    if not output_path.exists():
        return output_path

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_dir / f"{source_path.stem}_matched_{stamp}.xlsx"


def _rich_text_segments(
    dna_items: list[str], rna_items: list[str],
) -> str | CellRichText:
    dna_segment = "/".join(dna_items)
    rna_segment = "/".join(rna_items)

    parts: list[str | TextBlock] = []
    if dna_segment:
        parts.append(TextBlock(InlineFont(color=DNA_COLOR), dna_segment))
    if rna_segment:
        prefix = "/" if dna_segment else ""
        parts.append(TextBlock(InlineFont(color=RNA_COLOR), f"{prefix}{rna_segment}"))

    if not parts:
        return ""

    return CellRichText(*parts)


def _rich_text_for_name(cell: MatchCell) -> str | CellRichText:
    if cell.text in {"No match", "Invalid Feature"}:
        return cell.text
    return _rich_text_segments(cell.dna_names, cell.rna_names) or cell.text


def _rich_text_for_formula(cell: MatchCell) -> str | CellRichText:
    if not cell.formula_text:
        return ""
    dna_non_empty = [f for f in cell.dna_formulas if f]
    rna_non_empty = [f for f in cell.rna_formulas if f]
    return _rich_text_segments(dna_non_empty, rna_non_empty) or cell.formula_text


def export_results(
    dataset: pd.DataFrame,
    match_cells: list[MatchCell],
    source_path: Path,
    output_dir: Path,
    formula_column_name: str,
    name_column_name: str,
) -> Path:
    return export_workbook_results(
        datasets={"Sheet1": dataset},
        match_cells_by_sheet={"Sheet1": match_cells},
        source_path=source_path,
        output_dir=output_dir,
        formula_column_name=formula_column_name,
        name_column_name=name_column_name,
    )


def export_workbook_results(
    datasets: dict[str, pd.DataFrame],
    match_cells_by_sheet: dict[str, list[MatchCell]],
    source_path: Path,
    output_dir: Path,
    formula_column_name: str,
    name_column_name: str,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = _build_output_path(source_path, output_dir)
    result_tables: dict[str, pd.DataFrame] = {}

    with pd.ExcelWriter(output_path) as writer:
        for sheet_name, dataset in datasets.items():
            match_cells = match_cells_by_sheet[sheet_name]
            result = dataset.copy()
            result[formula_column_name] = [cell.formula_text for cell in match_cells]
            result[name_column_name] = [cell.text for cell in match_cells]
            result.to_excel(writer, sheet_name=sheet_name, index=False)
            result_tables[sheet_name] = result

    workbook = load_workbook(output_path)
    for sheet_name, result in result_tables.items():
        sheet = workbook[sheet_name]
        formula_col_idx = result.columns.get_loc(formula_column_name) + 1
        name_col_idx = result.columns.get_loc(name_column_name) + 1
        for row_index, match_cell in enumerate(match_cells_by_sheet[sheet_name], start=2):
            sheet.cell(row=row_index, column=formula_col_idx).value = _rich_text_for_formula(match_cell)
            sheet.cell(row=row_index, column=name_col_idx).value = _rich_text_for_name(match_cell)

    workbook.save(output_path)
    return output_path
