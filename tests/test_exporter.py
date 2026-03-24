import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

from ms_feature_db_matcher.exporter import export_results
from ms_feature_db_matcher.matcher import MatchCell


def test_export_results_writes_formula_and_name_columns(tmp_path) -> None:
    source = tmp_path / "input.xlsx"
    pd.DataFrame({"Feature": ["268.1052/17.59(Mz/RT)"]}).to_excel(source, index=False)

    output = tmp_path / "Output"
    result_path = export_results(
        dataset=pd.read_excel(source),
        match_cells=[MatchCell(
            text="dX/m1A",
            formula_text="C10H13N4O4/C11H16N5O3",
            dna_names=["dX"], rna_names=["m1A"],
            dna_formulas=["C10H13N4O4"], rna_formulas=["C11H16N5O3"],
        )],
        source_path=source,
        output_dir=output,
        formula_column_name="Matched Formula",
        name_column_name="Matched Short name",
    )

    workbook = load_workbook(result_path)
    sheet = workbook.active

    assert result_path.parent == output
    assert sheet.cell(row=1, column=2).value == "Matched Formula"
    assert sheet.cell(row=1, column=3).value == "Matched Short name"
    assert sheet.cell(row=2, column=3).value == "dX/m1A"


def test_export_results_preserves_dna_and_rna_rich_text_colors(tmp_path) -> None:
    source = tmp_path / "input.xlsx"
    dataset = pd.DataFrame({"Feature": ["268.1052/17.59(Mz/RT)"]})

    result_path = export_results(
        dataset=dataset,
        match_cells=[MatchCell(
            text="dX/m1A",
            formula_text="C10H13N4O4/C11H16N5O3",
            dna_names=["dX"], rna_names=["m1A"],
            dna_formulas=["C10H13N4O4"], rna_formulas=["C11H16N5O3"],
        )],
        source_path=source,
        output_dir=tmp_path / "Output",
        formula_column_name="Matched Formula",
        name_column_name="Matched Short name",
    )

    workbook = load_workbook(result_path, rich_text=True)
    sheet = workbook.active

    name_value = sheet.cell(row=2, column=3).value
    assert isinstance(name_value, CellRichText)
    assert str(name_value) == "dX/m1A"
    assert name_value[0].text == "dX"
    assert name_value[0].font.color.rgb == "000000FF"
    assert name_value[1].text == "/m1A"
    assert name_value[1].font.color.rgb == "00FF0000"

    formula_value = sheet.cell(row=2, column=2).value
    assert isinstance(formula_value, CellRichText)
    assert str(formula_value) == "C10H13N4O4/C11H16N5O3"
    assert formula_value[0].font.color.rgb == "000000FF"
    assert formula_value[1].font.color.rgb == "00FF0000"


def test_export_results_formula_empty_when_no_formula(tmp_path) -> None:
    source = tmp_path / "input.xlsx"
    dataset = pd.DataFrame({"Feature": ["268.1052/17.59(Mz/RT)"]})

    result_path = export_results(
        dataset=dataset,
        match_cells=[MatchCell(
            text="dX",
            formula_text="",
            dna_names=["dX"], rna_names=[],
            dna_formulas=[], rna_formulas=[],
        )],
        source_path=source,
        output_dir=tmp_path / "Output",
        formula_column_name="Matched Formula",
        name_column_name="Matched Short name",
    )

    workbook = load_workbook(result_path)
    sheet = workbook.active

    assert not sheet.cell(row=2, column=2).value  # empty formula
    assert sheet.cell(row=2, column=3).value == "dX"
